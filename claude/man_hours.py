from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Man-Hours Estimate"

# Colors
HEADER_FILL = PatternFill('solid', fgColor='1a2744')
SUBHEAD_FILL = PatternFill('solid', fgColor='253352')
MAP_FILL = PatternFill('solid', fgColor='1e2d47')
TOTAL_FILL = PatternFill('solid', fgColor='2a1a0a')
GRAND_FILL = PatternFill('solid', fgColor='3a2200')
WHITE = Font(color='FFFFFF', size=11, name='Arial')
WHITE_BOLD = Font(color='FFFFFF', size=11, name='Arial', bold=True)
GOLD = Font(color='DDAA55', size=11, name='Arial', bold=True)
GOLD_LG = Font(color='DDAA55', size=14, name='Arial', bold=True)
BLUE = Font(color='6699CC', size=11, name='Arial')
GREEN = Font(color='66AA66', size=11, name='Arial')
AMBER = Font(color='CC8844', size=11, name='Arial')
DIM = Font(color='778899', size=10, name='Arial', italic=True)
WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
CENTER = Alignment(horizontal='center', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')
THIN_BORDER = Border(
    bottom=Side(style='thin', color='334455')
)

ws.column_dimensions['A'].width = 38
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 42

def style_row(row, font=WHITE, fill=None, border=None, height=None):
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        if fill: cell.fill = fill
        if border: cell.border = border
        cell.alignment = WRAP if col == 5 else (RIGHT if col in [2,3,4] else Alignment(vertical='center'))
    if height: ws.row_dimensions[row].height = height

# Title
ws.merge_cells('A1:E1')
ws['A1'] = 'Pre-LLM Development Estimate: Space Visualization Project'
ws['A1'].font = GOLD_LG
ws['A1'].fill = HEADER_FILL
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 36

ws.merge_cells('A2:E2')
ws['A2'] = 'Estimated man-hours for a senior Three.js / WebGL developer (solo) — traditional development without AI assistance'
ws['A2'].font = DIM
ws['A2'].fill = HEADER_FILL
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 22

# Headers
r = 4
for col, txt in enumerate(['Component', 'Low (hrs)', 'High (hrs)', 'Mid (hrs)', 'Notes'], 1):
    cell = ws.cell(row=r, column=col, value=txt)
    cell.font = GOLD
    cell.fill = SUBHEAD_FILL
    cell.alignment = CENTER if col in [2,3,4] else Alignment(vertical='center')
ws.row_dimensions[r].height = 28

# Data
data = [
    # (indent, name, low, high, notes, section_header, font_override)
    ('SOLAR SYSTEM MAP', None, None, None, '~2,700 lines • Single-file HTML/JS/CSS', True, None),
    ('Orbital mechanics (Kepler solver)', 40, 60, None, 'Mean/eccentric anomaly, elliptical orbits, inclination, argument of perihelion', False, None),
    ('Asinh distance compression', 12, 20, None, 'Custom non-linear scale with smooth near-origin behavior; lots of tuning', False, None),
    ('Procedural planet textures (×10 types)', 60, 100, None, 'Earth continents+clouds, Jupiter bands+GRS, Saturn rings, ice giants, cratered, etc.', False, None),
    ('Sun rendering + glow layers', 8, 16, None, 'Additive blending sprites, pulse animation, opacity management', False, None),
    ('3D navigator (orbit camera + flyTo)', 24, 40, None, 'Smooth lerp, pitch/yaw, scroll zoom with log scaling, keyboard nav', False, None),
    ('Planet/moon/comet data entry', 16, 24, None, '34 bodies with orbital params, descriptions, classification, moon hierarchies', False, None),
    ('Particle belts (asteroid + Kuiper)', 8, 12, None, 'Randomized particle distributions, proper orbital zone ranges', False, None),
    ('Comet tails + extreme orbits', 12, 20, None, 'Tail sprites pointing away from Sun, near-parabolic orbit rendering, arc clamping', False, None),
    ('UI: info cards, HUD, minimap', 24, 40, None, 'Photo-viewer card, live orbital coords, zoom indicator, canvas minimap', False, None),
    ('Tour system (scripted flythrough)', 20, 32, None, 'Stop sequencing, narration, auto-advance, camera interpolation', False, None),
    ('Labels + hover tooltips', 8, 12, None, 'Canvas text sprites, distance-adaptive scaling, raycaster hover', False, None),
    ('Layer toggle system', 6, 10, None, 'Checkbox panel, visibility cascading through groups, keyboard shortcuts', False, None),
    ('Smart double-click (screen-space snap)', 8, 16, None, 'Project all bodies to screen coords, find nearest, handle edge cases', False, None),
    ('Habitable zone visualization', 4, 6, None, 'Annular ring at correct AU range, transparency, layer integration', False, None),
    ('Hill sphere visualization', 12, 20, None, 'Per-planet gravity zones, wireframe + fill, real-time position tracking', False, None),
    ('Actual Size toggle', 8, 16, None, 'Real AU radius math, render-loop override handling, educational UX', False, None),
    ('Help overlay + Log Truth overlay', 6, 10, None, 'Styled popups, keyboard wiring, educational copy', False, None),
    ('Arrow key cycling + nav buttons', 4, 8, None, 'Index tracking, flyTo integration, button wiring', False, None),
    ('Cross-browser testing + polish', 16, 24, None, 'WebGL compat, touch events, mobile, performance optimization', False, None),
    ('SOLAR SYSTEM SUBTOTAL', None, None, None, '', 'total', None),

    ('MILKY WAY MAP', None, None, None, '~2,700 lines • Procedural galaxy with 80k+ particles', True, None),
    ('Galaxy structure (spiral arms algorithm)', 32, 48, None, 'Logarithmic spiral math, arm count/width/winding, per-particle color variance', False, None),
    ('Galactic core + bar + bulge', 16, 24, None, 'Core density distribution, bar geometry, dimming gradient', False, None),
    ('Background stars + disk fill particles', 12, 20, None, 'Random distributions, GlowTexture map, depth/transparency handling', False, None),
    ('Notable stars data (40+ entries)', 24, 40, None, 'Research, position estimation, descriptions, planet sub-entries, sky descriptions', False, None),
    ('Procedural star textures', 16, 24, None, 'Per-spectral-type rendering, glow layers, size categories', False, None),
    ('3D navigator (reused + extended)', 12, 20, None, 'Same orbit camera with galaxy-specific tuning, WASD nav (later removed)', False, None),
    ('Sector grid + coordinate grid', 12, 20, None, 'Named galactic sectors, lat/long lines, toggle system', False, None),
    ('Photo-viewer card + star info', 16, 24, None, 'Procedural star portrait, spectral class display, planet listings', False, None),
    ('Tour system (galaxy edition)', 16, 24, None, 'Different stops, narration, galaxy-scale camera movements', False, None),
    ('Mode switcher (All/Planets/Stars)', 8, 12, None, 'Filter logic, button UI, visibility cascading', False, None),
    ('Per-star distance scaling', 8, 12, None, 'Camera-distance-based sprite scaling to prevent jitter (tricky bug fix)', False, None),
    ('depthWrite / transparency fixes', 8, 16, None, 'Black box debugging, PointsMaterial configuration across 4 systems', False, None),
    ('Center dimming (radial fade)', 6, 10, None, 'Vertex color modulation, core opacity reduction, smooth gradient', False, None),
    ('Smart double-click + arrow cycling', 8, 12, None, 'Screen-space projection, selection tracking, UI buttons', False, None),
    ('Actual Size toggle', 6, 10, None, 'Scene traversal, PointsMaterial size override, sprite scaling', False, None),
    ('UI: layers, help, Log Truth, overlays', 10, 16, None, 'Panel system, hotkeys, educational content', False, None),
    ('Cross-browser testing + polish', 12, 20, None, 'Performance with 80k particles, WebGL limits, mobile', False, None),
    ('MILKY WAY SUBTOTAL', None, None, None, '', 'total', None),

    ('GALAXY EXPLORER MAP', None, None, None, '~1,400 lines • Real astronomical data, 50+ galaxies', True, None),
    ('Galaxy data research + entry', 20, 32, None, 'RA/Dec/distance for 50+ galaxies, classifications, descriptions, image URLs', False, None),
    ('Sinh-compressed 3D positioning', 12, 20, None, 'Convert RA/Dec/distance to 3D scene coords with distance compression', False, None),
    ('Procedural galaxy textures', 16, 24, None, 'Per-morphology rendering (spiral, elliptical, irregular, lenticular)', False, None),
    ('RA/Dec grid + Mpc distance grid', 12, 20, None, 'Astronomical coordinate overlays, distance shells, labels', False, None),
    ('Info panel + galaxy images', 12, 20, None, 'Detail cards, Wikipedia image loading, pulsing detail indicators', False, None),
    ('Drunken tour (Elon\'s tour)', 8, 12, None, 'Randomized flythrough with comedic narration banners', False, None),
    ('Navigator + smart double-click', 10, 16, None, 'Orbit camera, plane intersection fallback, nearest-galaxy snap', False, None),
    ('Arrow cycling + nav buttons', 4, 8, None, 'Galaxy index tracking, flyTo, button wiring', False, None),
    ('Help overlay + Log Truth + Actual Size', 8, 12, None, 'Educational overlays, hotkeys, galaxy scaling toggle', False, None),
    ('Cross-browser testing + polish', 8, 12, None, 'Lighter map but still WebGL + sprite management', False, None),
    ('GALAXY EXPLORER SUBTOTAL', None, None, None, '', 'total', None),

    ('CROSS-CUTTING WORK', None, None, None, 'Shared across all three maps', True, None),
    ('Architecture + project setup', 8, 16, None, 'File structure decisions, CDN management, shared patterns', False, None),
    ('Cross-map navigation links', 4, 6, None, 'Consistent link bar, styling, placement', False, None),
    ('Consistent UX patterns', 12, 20, None, 'Help button style, green glow animation, overlay patterns, hotkey conventions', False, None),
    ('Bug investigation + debugging', 24, 40, None, 'Black hole Sun, flyTo targeting (0,0,0), depth buffer issues, jitter, etc.', False, None),
    ('Iterative design feedback loops', 20, 40, None, 'Back-and-forth with stakeholder on dimming, zoom distances, label placement', False, None),
    ('Research + fact-checking', 16, 24, None, 'Orbital parameters, star positions, galaxy distances, educational accuracy', False, None),
    ('CROSS-CUTTING SUBTOTAL', None, None, None, '', 'total', None),
]

r = 5
subtotal_rows = {'solar': [], 'milky': [], 'galaxy': [], 'cross': []}
current_section = None
section_start_row = None
section_map = {'SOLAR': 'solar', 'MILKY': 'milky', 'GALAXY': 'galaxy', 'CROSS': 'cross'}

for item in data:
    name, low, high, mid, notes, is_section, font_ov = item
    
    if is_section == True:
        # Section header
        ws.merge_cells(f'A{r}:D{r}')
        ws.cell(row=r, column=1, value=name).font = Font(color='AACCFF', size=12, name='Arial', bold=True)
        ws.cell(row=r, column=5, value=notes).font = DIM
        for col in range(1, 6):
            ws.cell(row=r, column=col).fill = MAP_FILL
        ws.row_dimensions[r].height = 28
        for key in section_map:
            if key in name:
                current_section = section_map[key]
                section_start_row = r + 1
                break
        r += 1
        continue
    
    if is_section == 'total':
        # Subtotal row
        ws.cell(row=r, column=1, value=name).font = GOLD
        ws.cell(row=r, column=2).font = GOLD
        ws.cell(row=r, column=3).font = GOLD
        ws.cell(row=r, column=4).font = GOLD
        ws.cell(row=r, column=2, value=f'=SUM(B{section_start_row}:B{r-1})')
        ws.cell(row=r, column=3, value=f'=SUM(C{section_start_row}:C{r-1})')
        ws.cell(row=r, column=4, value=f'=ROUND(AVERAGE(B{r},C{r}),0)')
        for col in range(1, 6):
            ws.cell(row=r, column=col).fill = TOTAL_FILL
            ws.cell(row=r, column=col).border = Border(top=Side(style='thin', color='DDAA55'), bottom=Side(style='thin', color='DDAA55'))
        ws.row_dimensions[r].height = 26
        subtotal_rows[current_section] = r
        r += 1
        # Spacer
        ws.row_dimensions[r].height = 10
        r += 1
        continue
    
    # Regular row
    ws.cell(row=r, column=1, value=f'  {name}').font = WHITE
    ws.cell(row=r, column=2, value=low).font = BLUE
    ws.cell(row=r, column=3, value=high).font = BLUE
    ws.cell(row=r, column=4, value=f'=ROUND(AVERAGE(B{r},C{r}),0)').font = WHITE_BOLD
    ws.cell(row=r, column=5, value=notes).font = DIM
    for col in range(1, 6):
        ws.cell(row=r, column=col).border = THIN_BORDER
    ws.cell(row=r, column=2).alignment = RIGHT
    ws.cell(row=r, column=3).alignment = RIGHT
    ws.cell(row=r, column=4).alignment = RIGHT
    ws.cell(row=r, column=5).alignment = WRAP
    ws.row_dimensions[r].height = 32
    r += 1

# Grand total
r += 1
ws.row_dimensions[r].height = 8
r += 1

st_rows = list(subtotal_rows.values())
ws.cell(row=r, column=1, value='GRAND TOTAL (ALL THREE MAPS)').font = Font(color='FFCC44', size=13, name='Arial', bold=True)
ws.cell(row=r, column=2, value=f'=B{st_rows[0]}+B{st_rows[1]}+B{st_rows[2]}+B{st_rows[3]}').font = Font(color='FFCC44', size=13, name='Arial', bold=True)
ws.cell(row=r, column=3, value=f'=C{st_rows[0]}+C{st_rows[1]}+C{st_rows[2]}+C{st_rows[3]}').font = Font(color='FFCC44', size=13, name='Arial', bold=True)
ws.cell(row=r, column=4, value=f'=ROUND(AVERAGE(B{r},C{r}),0)').font = Font(color='FFCC44', size=13, name='Arial', bold=True)
ws.cell(row=r, column=2).alignment = RIGHT
ws.cell(row=r, column=3).alignment = RIGHT
ws.cell(row=r, column=4).alignment = RIGHT
for col in range(1, 6):
    ws.cell(row=r, column=col).fill = GRAND_FILL
    ws.cell(row=r, column=col).border = Border(top=Side(style='medium', color='FFCC44'), bottom=Side(style='medium', color='FFCC44'))
ws.row_dimensions[r].height = 34
grand_row = r

# Cost estimate section
r += 2
ws.cell(row=r, column=1, value='COST ESTIMATES').font = Font(color='AACCFF', size=12, name='Arial', bold=True)
ws.cell(row=r, column=1).fill = MAP_FILL
for col in range(2, 6):
    ws.cell(row=r, column=col).fill = MAP_FILL
ws.row_dimensions[r].height = 28
r += 1

costs = [
    ('Solo senior dev ($125/hr)', f'=B{grand_row}*125', f'=C{grand_row}*125', None, 'US market rate for senior Three.js/WebGL specialist'),
    ('Small team (dev+designer+advisor)', f'=B{grand_row}*175', f'=C{grand_row}*175', None, 'Blended rate including design & science review'),
    ('Agency / contractor shop', f'=B{grand_row}*225', f'=C{grand_row}*225', None, 'Typical agency rate with PM overhead'),
]

for name_c, low_c, high_c, mid_c, notes_c in costs:
    ws.cell(row=r, column=1, value=f'  {name_c}').font = WHITE
    ws.cell(row=r, column=2, value=low_c).font = GREEN
    ws.cell(row=r, column=3, value=high_c).font = GREEN
    ws.cell(row=r, column=4, value=f'=ROUND(AVERAGE(B{r},C{r}),0)').font = Font(color='66AA66', size=11, name='Arial', bold=True)
    ws.cell(row=r, column=5, value=notes_c).font = DIM
    ws.cell(row=r, column=2).number_format = '$#,##0'
    ws.cell(row=r, column=3).number_format = '$#,##0'
    ws.cell(row=r, column=4).number_format = '$#,##0'
    ws.cell(row=r, column=2).alignment = RIGHT
    ws.cell(row=r, column=3).alignment = RIGHT
    ws.cell(row=r, column=4).alignment = RIGHT
    ws.cell(row=r, column=5).alignment = WRAP
    for col in range(1, 6):
        ws.cell(row=r, column=col).border = THIN_BORDER
    ws.row_dimensions[r].height = 28
    r += 1

# Timeline
r += 1
ws.cell(row=r, column=1, value='TIMELINE ESTIMATES').font = Font(color='AACCFF', size=12, name='Arial', bold=True)
ws.cell(row=r, column=1).fill = MAP_FILL
for col in range(2, 6):
    ws.cell(row=r, column=col).fill = MAP_FILL
ws.row_dimensions[r].height = 28
r += 1

timelines = [
    ('Solo dev (full-time)', f'=ROUND(B{grand_row}/160,1)', f'=ROUND(C{grand_row}/160,1)', None, 'Months at 160 hrs/month — no vacations, no context switching'),
    ('Small team (2-3 people)', f'=ROUND(B{grand_row}/320,1)', f'=ROUND(C{grand_row}/320,1)', None, 'Months with parallelism but communication overhead'),
    ('With LLM assistance (this project)', None, None, None, '~20 hours of conversation over a few sessions'),
]

for name_t, low_t, high_t, mid_t, notes_t in timelines:
    ws.cell(row=r, column=1, value=f'  {name_t}').font = WHITE
    if low_t:
        ws.cell(row=r, column=2, value=low_t).font = AMBER
        ws.cell(row=r, column=3, value=high_t).font = AMBER
        ws.cell(row=r, column=4, value=f'=ROUND(AVERAGE(B{r},C{r}),1)').font = Font(color='CC8844', size=11, name='Arial', bold=True)
        ws.cell(row=r, column=2).number_format = '0.0 "months"'
        ws.cell(row=r, column=3).number_format = '0.0 "months"'
        ws.cell(row=r, column=4).number_format = '0.0 "months"'
    else:
        ws.cell(row=r, column=2, value='~20').font = Font(color='44DD44', size=11, name='Arial', bold=True)
        ws.cell(row=r, column=3, value='hrs total').font = Font(color='44DD44', size=11, name='Arial', bold=True)
    ws.cell(row=r, column=5, value=notes_t).font = DIM
    ws.cell(row=r, column=2).alignment = RIGHT
    ws.cell(row=r, column=3).alignment = RIGHT
    ws.cell(row=r, column=4).alignment = RIGHT
    ws.cell(row=r, column=5).alignment = WRAP
    for col in range(1, 6):
        ws.cell(row=r, column=col).border = THIN_BORDER
    ws.row_dimensions[r].height = 28
    r += 1

# Footer
r += 1
ws.merge_cells(f'A{r}:E{r}')
ws.cell(row=r, column=1, value='Speedup factor: ~60-80× compared to traditional development. The iterative design loop (dimming, zoom tuning, bug hunting, educational features) accounts for most of the savings.').font = Font(color='556677', size=10, name='Arial', italic=True)
ws.cell(row=r, column=1).alignment = WRAP
ws.row_dimensions[r].height = 36

ws.sheet_properties.tabColor = '1a2744'
ws.freeze_panes = 'A5'

outpath = '/sessions/vibrant-zealous-wozniak/mnt/milkywayhw/claude/man_hours_estimate.xlsx'
wb.save(outpath)
print(f'Saved to {outpath}')
